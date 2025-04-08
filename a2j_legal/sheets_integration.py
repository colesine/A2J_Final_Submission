"""
Google Sheets Integration Module

This module handles the integration with Google Sheets for exporting case data.
"""

import os
import logging
import pickle
from typing import Optional, Dict, Any

import openpyxl
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Configure logging
logger = logging.getLogger(__name__)

class SheetsManager:
    """Class for managing Google Sheets integration."""
    
    def __init__(self, client_secret_file: str, spreadsheet_id: str, target_sheet_title: str):
        """
        Initialize the SheetsManager.
        
        Args:
            client_secret_file: Path to the OAuth client secret JSON file
            spreadsheet_id: ID of the Google Spreadsheet
            target_sheet_title: Title of the target sheet in the spreadsheet
        """
        self.client_secret_file = client_secret_file
        self.spreadsheet_id = spreadsheet_id
        self.target_sheet_title = target_sheet_title
        self.scopes = ['https://www.googleapis.com/auth/spreadsheets']
    
    def get_credentials(self):
        """
        Get credentials from saved token file or run the OAuth2 flow.
        
        Returns:
            OAuth2 credentials
        """
        # Check for token.pickle file first
        token_path = os.path.join(os.path.dirname(os.path.abspath(self.client_secret_file)), 'token.pickle')
        if os.path.exists(token_path):
            try:
                with open(token_path, 'rb') as token:
                    logger.info(f"Using saved credentials from {token_path}")
                    return pickle.load(token)
            except Exception as e:
                logger.error(f"Error loading token from {token_path}: {e}")
        
        # Fall back to the regular flow if token doesn't exist or is invalid
        logger.info("No valid token found, running OAuth flow")
        flow = InstalledAppFlow.from_client_secrets_file(self.client_secret_file, self.scopes)
        creds = flow.run_local_server(port=0)
        
        # Save the credentials for future use
        try:
            with open(token_path, 'wb') as token:
                pickle.dump(creds, token)
            logger.info(f"Credentials saved to {token_path}")
        except Exception as e:
            logger.error(f"Error saving token to {token_path}: {e}")
            
        return creds
    
    def hex_to_rgb(self, rgb_str: str) -> Optional[Dict[str, float]]:
        """
        Convert an ARGB hex string (e.g. "FF00FF00") to a dictionary with keys "red", "green" and "blue"
        on a 0-1 scale. The first two characters (alpha) are ignored.
        
        Args:
            rgb_str: ARGB hex string
            
        Returns:
            Dictionary with RGB values on a 0-1 scale, or None if conversion fails
        """
        if not rgb_str or len(rgb_str) < 8:
            return None
        try:
            r = int(rgb_str[2:4], 16) / 255.0
            g = int(rgb_str[4:6], 16) / 255.0
            b = int(rgb_str[6:8], 16) / 255.0
            return {"red": r, "green": g, "blue": b}
        except Exception:
            return None
    
    def get_sheet_id(self, service, spreadsheet_id: str, sheet_title: str) -> int:
        """
        Retrieve the internal sheetId for a given sheet title.
        
        Args:
            service: Google Sheets API service
            spreadsheet_id: ID of the Google Spreadsheet
            sheet_title: Title of the sheet
            
        Returns:
            The sheet ID
            
        Raises:
            Exception: If the sheet with the given title is not found
        """
        spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheets = spreadsheet.get('sheets', [])
        for sheet in sheets:
            props = sheet.get('properties', {})
            if props.get('title') == sheet_title:
                return props.get('sheetId')
        raise Exception(f"Sheet with title '{sheet_title}' not found.")
    
    def export_excel_to_sheets(self, excel_path: str) -> bool:
        """
        Export an Excel file to Google Sheets.
        
        Args:
            excel_path: Path to the Excel file
            
        Returns:
            True if the export was successful, False otherwise
        """
        try:
            # Get credentials and build service
            creds = self.get_credentials()
            service = build('sheets', 'v4', credentials=creds)
            
            # Load Excel file
            wb = openpyxl.load_workbook(excel_path, data_only=False)
            ws = wb.active
            
            # Get sheet ID
            sheet_id = self.get_sheet_id(service, self.spreadsheet_id, self.target_sheet_title)
            
            # Build a grid of CellData objects
            cell_data_rows = []
            
            # Process all rows including header
            for row in ws.iter_rows():
                row_values = []
                for cell in row:
                    cell_data = {}
                    # Retrieve the cell text; if None, use an empty string
                    cell_text = "" if cell.value is None else str(cell.value)
                    
                    # If the cell contains a hyperlink, embed it as a clickable text
                    if cell.hyperlink:
                        display_text = cell_text if cell_text != "" else cell.hyperlink.target
                        # Set the cell's value as a plain string
                        cell_data["userEnteredValue"] = {"stringValue": display_text}
                        # Build a text format run for the hyperlink
                        run_format = {"link": {"uri": cell.hyperlink.target}}
                        if cell.font:
                            if cell.font.bold:
                                run_format["bold"] = True
                            if cell.font.italic:
                                run_format["italic"] = True
                            if cell.font.underline:
                                run_format["underline"] = True
                        # Apply the text format run spanning the entire text
                        cell_data["textFormatRuns"] = [
                            {"startIndex": 0, "format": run_format}
                        ]
                    else:
                        # No hyperlink: set the value and apply any text formatting
                        if isinstance(cell.value, (int, float)):
                            cell_data["userEnteredValue"] = {"numberValue": cell.value}
                        else:
                            cell_data["userEnteredValue"] = {"stringValue": cell_text}
                        if cell.font:
                            text_format = {}
                            if cell.font.bold:
                                text_format["bold"] = True
                            if cell.font.italic:
                                text_format["italic"] = True
                            if cell.font.underline:
                                text_format["underline"] = True
                            if text_format:
                                cell_data.setdefault("userEnteredFormat", {})["textFormat"] = text_format

                    # Handle background fill; ignore default black or white fills
                    bg_color = None
                    if cell.fill and cell.fill.patternType == "solid":
                        if cell.fill.fgColor and cell.fill.fgColor.type == 'rgb' and cell.fill.fgColor.rgb:
                            color_hex = cell.fill.fgColor.rgb
                            if color_hex not in ("FF000000", "00000000", "FFFFFFFF", "00FFFFFF"):
                                bg_color = self.hex_to_rgb(color_hex)
                    if bg_color:
                        cell_data.setdefault("userEnteredFormat", {})["backgroundColor"] = bg_color

                    # If the cell has a comment (e.g. for red flagged cells), add it as a note
                    if cell.comment is not None:
                        cell_data["note"] = cell.comment.text

                    row_values.append(cell_data)
                
                cell_data_rows.append({"values": row_values})

            num_rows = len(cell_data_rows)
            num_cols = max(len(row["values"]) for row in cell_data_rows) if num_rows > 0 else 0

            update_request = {
                "updateCells": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": num_rows,
                        "startColumnIndex": 0,
                        "endColumnIndex": num_cols
                    },
                    "rows": cell_data_rows,
                    "fields": "*"
                }
            }

            # Clear the target sheet before updating
            try:
                clear_body = {}
                service.spreadsheets().values().clear(
                    spreadsheetId=self.spreadsheet_id,
                    range=self.target_sheet_title,
                    body=clear_body
                ).execute()
            except HttpError as error:
                logger.error(f"Error clearing the sheet: {error}")
                return False

            # Execute the batchUpdate to update cell values, formatting and notes
            try:
                body = {"requests": [update_request]}
                service.spreadsheets().batchUpdate(
                    spreadsheetId=self.spreadsheet_id,
                    body=body
                ).execute()
                sheet_url = f"https://docs.google.com/spreadsheets/d/{self.spreadsheet_id}/edit#gid={sheet_id}"
                logger.info(f"Sheet updated successfully. You can access the updated Google Sheet at: {sheet_url}")
                logger.info(f"Exported {ws.max_row} rows (including header) with all formatting preserved")
                return True
            except HttpError as error:
                logger.error(f"Error updating cells: {error}")
                return False
                
        except Exception as e:
            logger.error(f"Error exporting to Google Sheets: {e}")
            return False
