"""
Excel Utilities Module

This module handles Excel file operations for storing and retrieving case data.
"""

import os
import re
import logging
import unicodedata
import urllib.parse
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional, Set

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy

from a2j_legal.scraper import Case

# Configure logging
logger = logging.getLogger(__name__)

class ExcelManager:
    """Class for managing Excel file operations."""
    
    def __init__(self, case_archive_folder: str = "case_archive"):
        """
        Initialize the ExcelManager.
        
        Args:
            case_archive_folder: The folder where case archives are stored
        """
        self.case_archive_folder = case_archive_folder
        
        # Create case archive folder if it doesn't exist
        os.makedirs(self.case_archive_folder, exist_ok=True)
        
        # Define column headers
        self.csv_header = [
            "Case Name",
            "Unique Name",
            "Citation",
            "Date of Judgment",
            # Prompt 1 answers
            "Length of marriage till IJ (include separation period)",
            "Length of marriage (exclude separation period)",
            "Number of children",
            "Wife's income (monthly)",
            "Husband's income (monthly)",
            "Single or dual income marriage",
            # Prompt 2 answers
            "Direct Contribution (Wife)",
            "Indirect Contribution (Wife)",
            "Average Ratio (Wife)",
            "Final Ratio",
            "Adjustments",
            "Adjustment Reason",
            "Custody Type"
        ]
        
        # Define highlight fill for mismatched cells
        self.highlight_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def find_latest_excel_file(self, exclude_path: Optional[str] = None) -> Optional[str]:
        """
        Find the most recent Excel file with the format DD_MM_YYYY.xlsx in the case_archive folder.
        
        Args:
            exclude_path: Path to exclude from the search (e.g., the current file being created)
            
        Returns:
            The path to the latest Excel file, or None if no file is found
        """
        excel_files = []
        file_pattern = re.compile(r'^\d{2}_\d{2}_\d{4}\.xlsx$')
        
        # Check case_archive folder
        if not os.path.exists(self.case_archive_folder):
            logger.info("Archive folder doesn't exist yet. No previous Excel files found.")
            return None
        
        # Get all Excel files matching the pattern from the archive folder
        for file in os.listdir(self.case_archive_folder):
            if file_pattern.match(file):
                file_path = os.path.join(self.case_archive_folder, file)
                
                # Skip the current file if it's the one we're excluding
                if exclude_path and os.path.abspath(file_path) == os.path.abspath(exclude_path):
                    continue
                    
                excel_files.append((file_path, os.path.getmtime(file_path)))
        
        if not excel_files:
            return None
        
        # Sort files by modification time (newest first)
        excel_files.sort(key=lambda x: x[1], reverse=True)
        return excel_files[0][0]
    
    def create_excel_with_headers(self, sheet_name: Optional[str] = None) -> Tuple[Workbook, Worksheet, str]:
        """
        Create a new Excel workbook with a single sheet and header row.
        Save the file in the case_archive directory with date format DD_MM_YYYY.xlsx.
        
        Args:
            sheet_name: Name of the worksheet (default is "Cases")
            
        Returns:
            A tuple of (Workbook, Worksheet, save_path)
        """
        # Get current date in DD_MM_YYYY format
        current_date = datetime.now().strftime("%d_%m_%Y")
        file_name = f"{current_date}.xlsx"
        
        # Create full save path
        save_path = os.path.join(self.case_archive_folder, file_name)
        
        # Set default sheet name if not provided
        if sheet_name is None:
            sheet_name = "Cases"
        
        # Create workbook and add headers
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(self.csv_header)  # Add the header row
        
        # Save workbook
        wb.save(save_path)
        logger.info(f"Excel file created at {save_path}")

        return wb, ws, save_path
    
    def clean_and_encode_text(self, text: str) -> str:
        """
        Clean and encode text for use in URLs.
        
        Args:
            text: The text to clean and encode
            
        Returns:
            The cleaned and encoded text
        """
        # Normalize Unicode characters
        text = unicodedata.normalize('NFKC', text)
        text = text.replace("–", "-").replace("—", "-").replace("‑", "-")  # dash variants → ASCII hyphen
        
        # URL-encode the cleaned text
        text = urllib.parse.quote(text, safe='%')
        text = text.replace('%27%27...%27', '&text=')
        text = text.replace('%22%22...%22', '&text=')
        text = text.replace('%22...%22', '&text=')
        text = text.replace('%27...%27', '&text=')
        text = text.replace('...', '&text=')
        text = text.replace('-', '%2D')
        
        return text
    
    def load_existing_cases(self, excel_path: str) -> Set[str]:
        """
        Load existing case names from Excel file.
        
        Args:
            excel_path: Path to the Excel file
            
        Returns:
            A set of unique case names
        """
        try:
            # Load with openpyxl to preserve formatting
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Get headers from worksheet
            headers = [cell.value for cell in ws[1]]
            header_map = {col: i for i, col in enumerate(headers) if col}
            
            # Extract unique names for checking duplicates
            unique_name_col = header_map.get("Unique Name", 1)  # Default to column B if not found
            
            # Start from row 2 (skip headers)
            existing_cases = set()
            for row in range(2, ws.max_row + 1):
                case_name = ws.cell(row=row, column=unique_name_col+1).value
                if case_name:
                    existing_cases.add(case_name.strip())
            
            logger.info(f"Loaded {len(existing_cases)} existing cases")
            return existing_cases
            
        except Exception as e:
            logger.error(f"Error loading existing data: {e}")
            return set()
    
    def process_and_save_cases(self, cases: List[Case], llm_results: Dict[str, Dict[str, Any]]) -> str:
        """
        Process cases with LLM results and save to Excel.
        
        Args:
            cases: List of Case objects
            llm_results: Dictionary of LLM results for each case
            
        Returns:
            Path to the saved Excel file
        """
        # Create new Excel file with headers
        wb, ws, save_path = self.create_excel_with_headers()
        
        # Find latest Excel file (excluding the one we just created) and load existing cases
        latest_excel = self.find_latest_excel_file(exclude_path=save_path)
        existing_cases = set()
        source_wb = None
        source_ws = None
        
        # Make sure we're not using the file we just created as the source
        if latest_excel and latest_excel != save_path:
            try:
                logger.info(f"Found existing data in {latest_excel} (will append after processing new cases)")
                # Load existing cases
                existing_cases = self.load_existing_cases(latest_excel)
                
                # Load workbook for copying later
                source_wb = openpyxl.load_workbook(latest_excel)
                source_ws = source_wb.active
                
            except Exception as e:
                logger.error(f"Error loading existing data: {e}")
        
        # Process new cases
        new_case_count = 0
        for case in cases:
            # Skip if the case already exists in our dataset
            if case.unique_title.strip() in existing_cases:
                logger.info(f"Skipping {case.title} as it already exists in the dataset")
                continue
            
            new_case_count += 1
            
            # Get LLM results for this case
            case_results = llm_results.get(case.unique_title, {})
            all_fields = case_results.get("all_fields", [])
            all_evidence = case_results.get("all_evidence", [])
            
            # Ensure we have enough fields for all columns
            while len(all_fields) < 13:  # Fill missing fields
                all_fields.append("NA")
            
            # Create the row data
            new_row = [case.title, case.unique_title, case.citation, case.date] + all_fields
            
            # Append the row to the worksheet
            ws.append(new_row)
            row_index = ws.max_row
            
            # Add hyperlinks to evidence
            for i, answer in enumerate(all_fields):
                col_index = 4 + i  # Offset for the first 4 columns
                cell = ws.cell(row=row_index, column=col_index + 1)
                
                # Hyperlink logic
                try:
                    if i < len(all_evidence):
                        evidence_line = all_evidence[i]
                        if evidence_line not in ["Not Discussed", "NA", "Undisclosed"] and isinstance(case.url, str):
                            short_snippet = self.clean_and_encode_text(evidence_line)
                            highlight_url = f"{case.url}#:~:text={short_snippet}"
                            cell.hyperlink = highlight_url
                            cell.style = "Hyperlink"
                except Exception as e:
                    logger.error(f"Error setting hyperlink for {case.title}, field {i}: {e}")
            
            # Handle mismatched cells
            try:
                differences = case_results.get("differences", [])
                gpt_fields = case_results.get("gpt_fields", [])
                gemini_fields = case_results.get("gemini_fields", [])
                
                column_map = {0: 10, 1: 13, 2: 14, 3: 15}  # Map of field indices to column indices
                
                mismatched_cells = [
                    (row_index, column_map[i]) for i, mismatch in enumerate(differences)
                    if mismatch and i in column_map
                ]
                
                # Highlight mismatched cells
                for row, col in mismatched_cells:
                    cell = ws.cell(row=row, column=col)
                    cell.fill = self.highlight_fill
                
                # Add comments to mismatched cells
                for i, mismatch in enumerate(differences):
                    if mismatch and i in column_map:
                        row = row_index
                        col = column_map[i]
                        cell = ws.cell(row=row, column=col)
                        
                        gpt_val = gpt_fields[i] if i < len(gpt_fields) else "NA"
                        gemini_val = gemini_fields[i] if i < len(gemini_fields) else "NA"
                        
                        comment_text = f"Mismatch:\nGPT: {gpt_val}\nGemini: {gemini_val}"
                        cell.comment = Comment(comment_text, "A2JBot")
                        
            except Exception as e:
                logger.error(f"Error handling mismatched cells for {case.title}: {e}")
            
            # Save periodically to prevent data loss
            if new_case_count % 5 == 0:
                try:
                    wb.save(save_path)
                    logger.info(f"Progress saved after {new_case_count} new cases")
                except Exception as save_err:
                    logger.error(f"Error saving progress: {save_err}")
        
        # After processing all new cases, append the existing rows with formatting
        if source_wb and source_ws:
            start_row = ws.max_row + 1
            logger.info(f"\nAppending {source_ws.max_row - 1} existing cases from previous file with formatting...")
            
            # Copy all rows except header row
            for src_row in range(2, source_ws.max_row + 1):
                dst_row = start_row + src_row - 2
                
                # Copy cell values and formatting
                for src_col in range(1, min(len(self.csv_header) + 1, source_ws.max_column + 1)):
                    src_cell = source_ws.cell(row=src_row, column=src_col)
                    dst_cell = ws.cell(row=dst_row, column=src_col, value=src_cell.value)
                    
                    # Copy formatting
                    if src_cell.has_style:
                        dst_cell.font = copy(src_cell.font)
                        dst_cell.border = copy(src_cell.border)
                        dst_cell.fill = copy(src_cell.fill)
                        dst_cell.number_format = src_cell.number_format
                        dst_cell.protection = copy(src_cell.protection)
                        dst_cell.alignment = copy(src_cell.alignment)
                    
                    # Copy hyperlink
                    if src_cell.hyperlink:
                        dst_cell.hyperlink = src_cell.hyperlink
                        dst_cell.style = "Hyperlink"
                    
                    # Copy comment if any
                    if src_cell.comment:
                        dst_cell.comment = Comment(src_cell.comment.text, src_cell.comment.author)
                
                if (src_row - 1) % 50 == 0:
                    logger.info(f"Appended {src_row - 1} of {source_ws.max_row - 1} existing rows with formatting")
            
            logger.info(f"Successfully appended {source_ws.max_row - 1} existing rows with formatting")
        
        try:
            # Apply column widths from source if available
            if source_ws:
                for col in range(1, min(len(self.csv_header) + 1, source_ws.max_column + 1)):
                    col_letter = get_column_letter(col)
                    if source_ws.column_dimensions[col_letter].width:
                        ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width
            
            wb.save(save_path)
            logger.info(f"✅ Excel file updated: {save_path}")
            logger.info(f"Process completed. Added {new_case_count} new cases and appended {source_ws.max_row - 1 if source_ws else 0} existing cases with formatting.")
            
            return save_path
            
        except Exception as final_save_err:
            logger.error(f"Error saving final file: {final_save_err}")
            return ""
